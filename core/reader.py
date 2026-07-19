import openpyxl
from copy import copy
from typing import List, Dict, Optional
from openpyxl.utils import get_column_letter as _get_column_letter


def _extract_cell_style(cell):
    """从源单元格提取样式信息（Font、Fill、Border、Alignment、number_format）。

    使用 copy.copy() 深拷贝样式对象，确保关闭源工作簿后样式仍然可用。
    对空单元格或无样式的单元格返回 None。
    """
    if cell is None or not cell.has_style:
        return None
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "number_format": cell.number_format,
    }


def _apply_style_to_cell(dst_cell, style):
    """将提取的样式字典应用到目标单元格。"""
    if style is None:
        return
    if style.get("font") is not None:
        dst_cell.font = style["font"]
    if style.get("fill") is not None:
        dst_cell.fill = style["fill"]
    if style.get("border") is not None:
        dst_cell.border = style["border"]
    if style.get("alignment") is not None:
        dst_cell.alignment = style["alignment"]
    if style.get("number_format") is not None and style["number_format"] != "General":
        dst_cell.number_format = style["number_format"]


def _extract_row_styles(row_cells):
    """从一行 Cell 对象中提取所有单元格的样式信息。"""
    return [_extract_cell_style(c) for c in row_cells]


def _copy_column_widths(src_ws, dst_ws):
    """从源工作表复制列宽到目标工作表。"""
    for col_letter, col_dim in src_ws.column_dimensions.items():
        if col_dim.width is not None:
            dst_ws.column_dimensions[col_letter].width = col_dim.width


def _copy_row_heights(src_ws, dst_ws, row_map):
    """根据行号映射复制行高。

    Args:
        src_ws: 源工作表
        dst_ws: 目标工作表
        row_map: {src_row: dst_row} 映射表
    """
    for src_row, dst_row in row_map.items():
        src_dim = src_ws.row_dimensions.get(src_row)
        if src_dim is not None and src_dim.height is not None:
            dst_ws.row_dimensions[dst_row].height = src_dim.height


def read_sheet_grouped_with_styles(
    file_path: str,
    sheet_name: str,
    column: str,
    header_row: int = 1,
    data_only: bool = True,
) -> tuple:
    """以普通模式读取工作表，按列分组，同时保留单元格样式。

    以 header_row（1-indexed）作为表头行。

    返回 (headers, header_styles, groups)：
    - headers: 表头值列表
    - header_styles: 表头样式列表（每个元素为样式字典或 None）
    - groups: {value: [(src_row, row_values, row_styles), ...]}
      其中 row_values 为值列表，row_styles 为样式字典列表
    """
    wb = openpyxl.load_workbook(file_path, data_only=data_only)
    ws = wb[sheet_name]

    all_rows = list(ws.iter_rows())
    if len(all_rows) < header_row:
        wb.close()
        return [], [], {}

    header_cells = all_rows[header_row - 1]
    headers = [c.value for c in header_cells]
    header_styles = _extract_row_styles(header_cells)

    headers_str = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(headers)]
    if column not in headers_str:
        wb.close()
        return headers, header_styles, {}

    col_idx = headers_str.index(column)

    groups: Dict[str, list] = {}
    for src_row, row_cells in enumerate(all_rows[header_row:], start=header_row + 1):
        key = row_cells[col_idx].value if col_idx < len(row_cells) else None
        key = "(空)" if key is None else str(key)
        row_values = [c.value for c in row_cells]
        row_styles = _extract_row_styles(row_cells)
        groups.setdefault(key, []).append((src_row, row_values, row_styles))

    wb.close()
    return headers, header_styles, groups


def get_sheet_names(file_path: str) -> List[str]:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def get_columns(file_path: str, sheet_name: str, header_row: int = 1) -> List[str]:
    """读取指定行作为列名（1-indexed，默认为第 1 行）。"""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[sheet_name]
    row = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    wb.close()
    return [str(c) if c is not None else f"Col{i}" for i, c in enumerate(row)]


def get_unique_values(file_path: str, sheet_name: str, column: str, header_row: int = 1) -> Dict[str, int]:
    """获取指定列的所有唯一值及其计数（1-indexed header_row）。"""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[sheet_name]
    header_iter = ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)
    try:
        headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(next(header_iter))]
    except StopIteration:
        wb.close()
        return {}
    if column not in headers:
        wb.close()
        return {}
    col_idx = headers.index(column)
    counts: Dict[str, int] = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if col_idx < len(row):
            val = row[col_idx]
            val = "(空)" if val is None else str(val)
        else:
            val = "(空)"
        counts[val] = counts.get(val, 0) + 1
    wb.close()
    return counts


def read_sheet_grouped(
    file_path: str,
    sheet_name: str,
    column: str,
    header_row: int = 1,
) -> tuple:
    """流式读取并按列分组。

    以 header_row（1-indexed）作为表头行，该行之前的所有行将被跳过，
    该行之后的所有行作为数据行。返回 (headers, {value: [row_tuple, ...]})。
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    row_iter = ws.iter_rows(values_only=True)

    # 跳过 header_row 之前的行
    for _ in range(header_row - 1):
        try:
            next(row_iter)
        except StopIteration:
            wb.close()
            return [], {}

    try:
        header_row_data = next(row_iter)
    except StopIteration:
        wb.close()
        return [], {}

    headers = [str(c) if c is not None else f"Col{i}" for i, c in enumerate(header_row_data)]
    if column not in headers:
        wb.close()
        return headers, {}
    col_idx = headers.index(column)

    groups: Dict[str, list] = {}
    for row in row_iter:
        key = row[col_idx] if col_idx < len(row) else None
        key = "(空)" if key is None else str(key)
        groups.setdefault(key, []).append(tuple(row))

    wb.close()
    return headers, groups


# ── 前置行读取 ──────────────────────────────────────────────

def read_leading_rows_with_styles(
    file_path: str,
    sheet_name: str,
    start_row: int,
    end_row: int,
    data_only: bool = True,
) -> list:
    """读取源文件中指定行范围的行及其样式。

    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称
        start_row: 起始行号（1-indexed，含）
        end_row: 结束行号（1-indexed，含）
        data_only: 是否只读数据（忽略公式计算值）

    Returns:
        [(src_row, row_values, row_styles), ...] 列表
    """
    if start_row < 1 or end_row < start_row:
        return []
    wb = openpyxl.load_workbook(file_path, data_only=data_only)
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows())
    if len(all_rows) < end_row:
        wb.close()
        return []
    leading = []
    for i, row_cells in enumerate(all_rows[start_row - 1:end_row], start=start_row):
        row_values = [c.value for c in row_cells]
        row_styles = _extract_row_styles(row_cells)
        leading.append((i, row_values, row_styles))
    wb.close()
    return leading


# ── 合并单元格复制 ──────────────────────────────────────────

def copy_merged_cells(
    src_ws,
    dst_ws,
    row_range_start: int,
    row_range_end: int,
    row_shift: int = 0,
):
    """将源工作表中指定行范围内的合并单元格复制到目标工作表。

    仅复制完全位于 [row_range_start, row_range_end] 内的合并区域。
    通过 row_shift 调整目标行号（正数下移，负数上移）。

    Args:
        src_ws: 源工作表
        dst_ws: 目标工作表
        row_range_start: 源行范围起始（1-indexed）
        row_range_end: 源行范围结束（1-indexed）
        row_shift: 目标行偏移量
    """
    for merged_range in src_ws.merged_cells.ranges:
        if merged_range.min_row < row_range_start or merged_range.max_row > row_range_end:
            continue
        new_min_row = merged_range.min_row + row_shift
        new_max_row = merged_range.max_row + row_shift
        try:
            start_cell = f"{_get_column_letter(merged_range.min_col)}{new_min_row}"
            end_cell = f"{_get_column_letter(merged_range.max_col)}{new_max_row}"
            dst_ws.merge_cells(f"{start_cell}:{end_cell}")
        except Exception:
            pass  # 合并区域冲突时静默跳过


def copy_merged_cells_with_map(
    src_ws,
    dst_ws,
    row_map: Dict[int, int],
):
    """根据行号映射将合并单元格复制到目标工作表。

    适用于数据行按分组重新排列后的合并单元格复制。

    Args:
        src_ws: 源工作表
        dst_ws: 目标工作表
        row_map: {src_row: dst_row} 映射表
    """
    if not row_map:
        return
    for merged_range in src_ws.merged_cells.ranges:
        src_min_row = merged_range.min_row
        src_max_row = merged_range.max_row
        if src_min_row not in row_map or src_max_row not in row_map:
            continue
        new_min_row = row_map[src_min_row]
        new_max_row = row_map[src_max_row]
        try:
            start_cell = f"{_get_column_letter(merged_range.min_col)}{new_min_row}"
            end_cell = f"{_get_column_letter(merged_range.max_col)}{new_max_row}"
            dst_ws.merge_cells(f"{start_cell}:{end_cell}")
        except Exception:
            pass


# ── 行范围预览 ──────────────────────────────────────────────

def read_row_range_preview(
    file_path: str,
    sheet_name: str,
    start_row: int,
    end_row: int,
    max_cells: int = 3,
) -> str:
    """读取指定行范围的前几列内容，用于 UI 预览。

    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名称
        start_row: 起始行号（1-indexed）
        end_row: 结束行号（1-indexed）
        max_cells: 每行最多显示几个列值

    Returns:
        内容摘要字符串，如 "第1行: Sales Report, 第2行: Q1 Data, ..."
        若读取失败或范围无效则返回空字符串。
    """
    if start_row < 1 or end_row < start_row:
        return ""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < end_row:
            wb.close()
            return ""
        parts = []
        for i in range(start_row - 1, end_row):
            row_data = all_rows[i]
            cells = [str(c) for c in row_data[:max_cells] if c is not None]
            row_label = f"Row{i + 1}"
            parts.append(f"{row_label}: {', '.join(cells)}" if cells else row_label)
        wb.close()
        return "; ".join(parts) if parts else ""
    except Exception:
        return ""
