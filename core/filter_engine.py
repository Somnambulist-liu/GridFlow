"""流式数据筛选引擎"""
import os
from openpyxl import load_workbook, Workbook
from PySide6.QtCore import QThread, Signal


class FilterWorker(QThread):
    progress = Signal(int, int, str)
    finished = Signal(str)
    error_occurred = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._file_path = ""
        self._sheet_name = ""
        self._conditions = []
        self._logic = "AND"
        self._output_dir = ""
        self._output_name = ""

    def configure(self, file_path: str, sheet_name: str, conditions: list,
                  logic: str = "AND", output_dir: str = "", output_name: str = "筛选结果.xlsx"):
        self._file_path = file_path
        self._sheet_name = sheet_name
        self._conditions = conditions
        self._logic = logic
        self._output_dir = output_dir
        self._output_name = output_name

    def run(self):
        try:
            wb = load_workbook(self._file_path, read_only=True)
            ws = wb[self._sheet_name]

            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            col_indices = {h: i for i, h in enumerate(headers)}

            # Verify condition columns exist
            for cond in self._conditions:
                if cond["column"] not in col_indices:
                    raise ValueError(f"列 '{cond['column']}' 不存在")

            # First pass: count matching rows
            self.progress.emit(0, 0, "正在扫描匹配行...")
            matching_rows = []
            total_rows = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                total_rows += 1
                if total_rows % 1000 == 0:
                    self.progress.emit(0, total_rows, f"已扫描 {total_rows} 行...")
                if self._evaluate_row(row, col_indices):
                    matching_rows.append(row)

            wb.close()

            if not matching_rows:
                self.finished.emit("没有匹配的数据行")
                return

            # Write output
            output_path = os.path.join(self._output_dir, self._output_name)
            out_wb = Workbook()
            out_ws = out_wb.active
            out_ws.append(headers)

            for i, row in enumerate(matching_rows):
                out_ws.append(list(row))
                if (i + 1) % 500 == 0:
                    self.progress.emit(i + 1, len(matching_rows), f"正在写入 {i + 1}/{len(matching_rows)} 行...")

            out_wb.save(output_path)
            out_wb.close()

            summary = (
                f"筛选完成！原始 {total_rows} 行 → 匹配 {len(matching_rows)} 行\n"
                f"输出文件：{os.path.basename(output_path)}"
            )
            self.progress.emit(len(matching_rows), len(matching_rows), "完成")
            self.finished.emit(summary)

        except Exception as e:
            self.error_occurred.emit(str(e))

    def _evaluate_row(self, row: tuple, col_indices: dict) -> bool:
        results = []
        for cond in self._conditions:
            col_idx = col_indices[cond["column"]]
            cell_value = row[col_idx]
            op = cond["operator"]
            target = cond.get("value")
            target2 = cond.get("value2")

            if op == "eq":
                results.append(str(cell_value) == str(target) if cell_value is not None else target == "")
            elif op == "neq":
                results.append(str(cell_value) != str(target) if cell_value is not None else target != "")
            elif op == "gt":
                try:
                    results.append(float(cell_value or 0) > float(target))
                except (ValueError, TypeError):
                    results.append(False)
            elif op == "lt":
                try:
                    results.append(float(cell_value or 0) < float(target))
                except (ValueError, TypeError):
                    results.append(False)
            elif op == "gte":
                try:
                    results.append(float(cell_value or 0) >= float(target))
                except (ValueError, TypeError):
                    results.append(False)
            elif op == "lte":
                try:
                    results.append(float(cell_value or 0) <= float(target))
                except (ValueError, TypeError):
                    results.append(False)
            elif op == "contains":
                results.append(target.lower() in str(cell_value or "").lower())
            elif op == "not_contains":
                results.append(target.lower() not in str(cell_value or "").lower())
            elif op == "between":
                try:
                    val = float(cell_value or 0)
                    results.append(float(target) <= val <= float(target2))
                except (ValueError, TypeError):
                    results.append(False)
            elif op == "is_empty":
                results.append(cell_value is None or str(cell_value).strip() == "")
            elif op == "not_empty":
                results.append(cell_value is not None and str(cell_value).strip() != "")

        if self._logic == "AND":
            return all(results)
        else:
            return any(results)
