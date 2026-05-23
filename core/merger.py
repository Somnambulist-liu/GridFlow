import os
import openpyxl
from PySide6.QtCore import QThread, Signal


class MergeWorker(QThread):
    progress = Signal(int, int, str)
    finished = Signal(str)
    error_occurred = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._is_cancelled = False

    def cancel(self):
        self._is_cancelled = True

    def configure(
        self,
        mode: str,                # "files" or "sheets"
        file_paths: list,         # for files mode: list of file paths
        file_path: str = "",      # for sheets mode: single file
        sheet_names: list = None, # for sheets mode: sheets to merge
        output_dir: str = "",
        output_name: str = "合并结果.xlsx",
    ):
        self.mode = mode
        self.file_paths = file_paths
        self.file_path = file_path
        self.sheet_names = sheet_names or []
        self.output_dir = output_dir
        self.output_name = output_name

    def run(self):
        try:
            if self.mode == "files":
                self._merge_files()
            else:
                self._merge_sheets()
        except Exception as e:
            self.error_occurred.emit(str(e))

    def _merge_files(self):
        total = len(self.file_paths)
        self.progress.emit(0, total, "开始合并...")
        all_rows = []
        headers = None

        for i, fp in enumerate(self.file_paths):
            if self._is_cancelled:
                return
            self.progress.emit(i + 1, total, f"读取：{os.path.basename(fp)}")
            wb = openpyxl.load_workbook(fp, read_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            try:
                hdr = [str(c) if c is not None else f"Col{j}" for j, c in enumerate(next(it))]
            except StopIteration:
                wb.close()
                continue
            if headers is None:
                headers = hdr
            for row in it:
                all_rows.append(tuple(row))
            wb.close()

        if headers is None:
            self.finished.emit("没有读取到任何数据")
            return

        self.progress.emit(total, total, "正在写入合并文件...")
        out_path = os.path.join(self.output_dir, self.output_name)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in all_rows:
            ws.append(row)
        wb.save(out_path)
        wb.close()

        summary = f"合并完成！共合并 {total} 个文件，{len(all_rows)} 行数据"
        self.progress.emit(total, total, summary)
        self.finished.emit(summary)

    def _merge_sheets(self):
        all_rows = []
        headers = None
        total = len(self.sheet_names)

        wb = openpyxl.load_workbook(self.file_path, read_only=True)
        for i, sn in enumerate(self.sheet_names):
            if self._is_cancelled:
                break
            self.progress.emit(i + 1, total, f"读取 Sheet：{sn}")
            ws = wb[sn]
            it = ws.iter_rows(values_only=True)
            try:
                hdr = [str(c) if c is not None else f"Col{j}" for j, c in enumerate(next(it))]
            except StopIteration:
                continue
            if headers is None:
                headers = hdr
            for row in it:
                all_rows.append(tuple(row))
        wb.close()

        if headers is None:
            self.finished.emit("没有读取到任何数据")
            return

        self.progress.emit(total, total, "正在写入合并文件...")
        out_path = os.path.join(self.output_dir, self.output_name)
        owb = openpyxl.Workbook()
        ows = owb.active
        ows.append(headers)
        for row in all_rows:
            ows.append(row)
        owb.save(out_path)
        owb.close()

        summary = f"合并完成！共合并 {total} 个 Sheet，{len(all_rows)} 行数据"
        self.progress.emit(total, total, summary)
        self.finished.emit(summary)
