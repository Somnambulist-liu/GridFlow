"""透视表引擎"""
import os
from openpyxl import load_workbook, Workbook
from PySide6.QtCore import QThread, Signal


class PivotWorker(QThread):
    progress = Signal(int, int, str)
    finished = Signal(str)
    error_occurred = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._file_path = ""
        self._sheet_name = ""
        self._row_field = ""
        self._col_field = ""
        self._value_field = ""
        self._agg_func = "count"
        self._output_dir = ""
        self._output_name = ""

    def configure(self, file_path: str, sheet_name: str,
                  row_field: str, col_field: str, value_field: str,
                  agg_func: str = "count", output_dir: str = "",
                  output_name: str = "透视表结果.xlsx"):
        self._file_path = file_path
        self._sheet_name = sheet_name
        self._row_field = row_field
        self._col_field = col_field
        self._value_field = value_field
        self._agg_func = agg_func
        self._output_dir = output_dir
        self._output_name = output_name

    def run(self):
        try:
            wb = load_workbook(self._file_path, read_only=True)
            ws = wb[self._sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if len(rows) < 2:
                self.finished.emit("数据不足（至少需要标题行 + 1 行数据）")
                return

            headers = list(rows[0])
            col_idx = {h: i for i, h in enumerate(headers)}

            for f in [self._row_field, self._col_field, self._value_field]:
                if f not in col_idx:
                    raise ValueError(f"列 '{f}' 不存在")

            row_idx = col_idx[self._row_field]
            col_id = col_idx[self._col_field]
            val_idx = col_idx[self._value_field]

            # Build pivot dict: {(row_val, col_val): [values]}
            pivot_data = {}
            total = len(rows) - 1
            for i, row in enumerate(rows[1:], 1):
                if i % 1000 == 0:
                    self.progress.emit(i, total, f"正在汇总 {i}/{total} 行...")
                rv = str(row[row_idx]) if row[row_idx] is not None else "(空)"
                cv = str(row[col_id]) if row[col_id] is not None else "(空)"
                v = row[val_idx]
                pivot_data.setdefault((rv, cv), []).append(v)

            # Compute aggregation
            row_vals = sorted(set(k[0] for k in pivot_data))
            col_vals = sorted(set(k[1] for k in pivot_data))

            agg_results = {}
            for (rv, cv), values in pivot_data.items():
                agg_results[(rv, cv)] = self._aggregate(values)

            # Write cross-tabulation
            output_path = os.path.join(self._output_dir, self._output_name)
            out_wb = Workbook()
            out_ws = out_wb.active

            header_row = [f"{self._row_field} \\ {self._col_field}"] + [str(c) for c in col_vals] + ["合计"]
            out_ws.append(header_row)

            for rv in row_vals:
                out_row = [str(rv)]
                row_total = 0
                for cv in col_vals:
                    val = agg_results.get((rv, cv), 0)
                    out_row.append(val)
                    row_total += val if isinstance(val, (int, float)) else 0
                out_row.append(row_total)
                out_ws.append(out_row)

            # Totals row
            total_row = ["合计"]
            for j, cv in enumerate(col_vals):
                col_total = sum(
                    agg_results.get((rv, cv), 0)
                    for rv in row_vals
                    if isinstance(agg_results.get((rv, cv), 0), (int, float))
                )
                total_row.append(col_total)
            grand_total = sum(v for v in total_row[1:] if isinstance(v, (int, float)))
            total_row.append(grand_total)
            out_ws.append(total_row)

            out_wb.save(output_path)
            out_wb.close()

            self.progress.emit(total, total, "完成")
            self.finished.emit(
                f"透视表生成完成！{len(row_vals)} 行 × {len(col_vals)} 列\n"
                f"输出文件：{os.path.basename(output_path)}"
            )

        except Exception as e:
            self.error_occurred.emit(str(e))

    def _aggregate(self, values: list):
        numeric = []
        for v in values:
            try:
                numeric.append(float(v))
            except (ValueError, TypeError):
                pass

        if self._agg_func == "count":
            return len(values)
        elif self._agg_func == "sum":
            return round(sum(numeric), 2)
        elif self._agg_func == "avg":
            return round(sum(numeric) / len(numeric), 2) if numeric else 0
        elif self._agg_func == "min":
            return min(numeric) if numeric else 0
        elif self._agg_func == "max":
            return max(numeric) if numeric else 0
        return len(values)
