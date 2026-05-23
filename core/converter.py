import os
import openpyxl
import csv
from PySide6.QtCore import QThread, Signal


class ConvertWorker(QThread):
    progress = Signal(int, int, str)
    finished = Signal(str)
    error_occurred = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._is_cancelled = False

    def cancel(self):
        self._is_cancelled = True

    def configure(
        self,
        file_paths: list,
        target_format: str,  # "xlsx" or "csv"
        output_dir: str,
    ):
        self.file_paths = file_paths
        self.target_format = target_format
        self.output_dir = output_dir

    def run(self):
        try:
            total = len(self.file_paths)
            ok = 0
            for i, fp in enumerate(self.file_paths):
                if self._is_cancelled:
                    break
                basename = os.path.basename(fp)
                self.progress.emit(i + 1, total, f"转换：{basename}")
                try:
                    if self.target_format == "csv":
                        self._xlsx_to_csv(fp)
                    else:
                        self._csv_to_xlsx(fp)
                    ok += 1
                except Exception as e:
                    self.error_occurred.emit(f"{basename} 转换失败：{e}")
            summary = f"转换完成！成功 {ok} / {total} 个文件"
            self.progress.emit(total, total, summary)
            self.finished.emit(summary)
        except Exception as e:
            self.error_occurred.emit(str(e))

    def _xlsx_to_csv(self, file_path: str):
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        base = os.path.splitext(os.path.basename(file_path))[0]
        out = os.path.join(self.output_dir, f"{base}.csv")
        with open(out, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)
        wb.close()

    def _csv_to_xlsx(self, file_path: str):
        base = os.path.splitext(os.path.basename(file_path))[0]
        out = os.path.join(self.output_dir, f"{base}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)
        wb.save(out)
        wb.close()
