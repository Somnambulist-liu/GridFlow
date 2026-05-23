"""数据校验引擎"""
import os
from openpyxl import load_workbook, Workbook
from PySide6.QtCore import QThread, Signal


class ValidateWorker(QThread):
    progress = Signal(int, int, str)
    finished = Signal(str)
    error_occurred = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._file_path = ""
        self._sheet_name = ""
        self._checks = {}
        self._output_dir = ""

    def configure(self, file_path: str, sheet_name: str, checks: dict, output_dir: str = ""):
        self._file_path = file_path
        self._sheet_name = sheet_name
        self._checks = checks
        self._output_dir = output_dir

    def run(self):
        try:
            wb = load_workbook(self._file_path, read_only=True)
            ws = wb[self._sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if len(rows) < 2:
                self.finished.emit("数据不足")
                return

            headers = list(rows[0])
            total_rows = len(rows) - 1
            data_rows = rows[1:]

            issues = {}
            report_lines = []

            # 1. Empty cell check
            if self._checks.get("empty", False):
                self.progress.emit(0, total_rows, "正在检测空值...")
                threshold = self._checks.get("empty_threshold", 50)
                empty_issues = []
                for col_idx, col_name in enumerate(headers):
                    empty_count = sum(1 for row in data_rows if row[col_idx] is None or str(row[col_idx]).strip() == "")
                    empty_pct = round(empty_count / total_rows * 100, 1) if total_rows > 0 else 0
                    if empty_pct >= threshold:
                        empty_issues.append(f"  {col_name}: {empty_count}/{total_rows} 行空值 ({empty_pct}%)")
                if empty_issues:
                    issues["空值检测"] = empty_issues
                    report_lines.append(f"[空值检测] 阈值 ≥{threshold}%: {len(empty_issues)} 列超标")
                    report_lines.extend(empty_issues)
                else:
                    report_lines.append(f"[空值检测] 所有列空值率 < {threshold}%，通过")

            # 2. Outlier check (IQR)
            if self._checks.get("outliers", False):
                self.progress.emit(0, total_rows, "正在检测异常值...")
                multiplier = self._checks.get("outlier_multiplier", 1.5)
                outlier_issues = []
                for col_idx, col_name in enumerate(headers):
                    values = []
                    for row in data_rows:
                        try:
                            values.append(float(row[col_idx]))
                        except (ValueError, TypeError):
                            pass
                    if len(values) < 4:
                        continue
                    values.sort()
                    n = len(values)
                    q1 = values[n // 4]
                    q3 = values[3 * n // 4]
                    iqr = q3 - q1
                    if iqr == 0:
                        continue
                    lower = q1 - multiplier * iqr
                    upper = q3 + multiplier * iqr
                    outliers = sum(1 for v in values if v < lower or v > upper)
                    if outliers > 0:
                        outlier_issues.append(
                            f"  {col_name}: {outliers} 个异常值 (范围 [{round(lower, 2)}, {round(upper, 2)}])"
                        )
                if outlier_issues:
                    issues["异常值检测"] = outlier_issues
                    report_lines.append(f"[异常值检测] IQR×{multiplier}: {len(outlier_issues)} 列存在异常值")
                    report_lines.extend(outlier_issues)
                else:
                    report_lines.append(f"[异常值检测] 未发现异常值，通过")

            # 3. Type consistency check
            if self._checks.get("type_check", False):
                self.progress.emit(0, total_rows, "正在检测类型一致性...")
                type_issues = []
                for col_idx, col_name in enumerate(headers):
                    types = set()
                    for row in data_rows:
                        v = row[col_idx]
                        if v is None or str(v).strip() == "":
                            types.add("empty")
                        else:
                            try:
                                float(v)
                                types.add("number")
                            except (ValueError, TypeError):
                                types.add("string")
                    if len(types) > 1:
                        type_issues.append(f"  {col_name}: 混合类型 {types}")
                if type_issues:
                    issues["类型检查"] = type_issues
                    report_lines.append(f"[类型检查] {len(type_issues)} 列类型不统一")
                    report_lines.extend(type_issues)
                else:
                    report_lines.append("[类型检查] 所有列类型一致，通过")

            # 4. Duplicate rows check
            if self._checks.get("duplicates", False):
                self.progress.emit(0, total_rows, "正在检测重复行...")
                seen = set()
                dup_count = 0
                for row in data_rows:
                    key = tuple(str(v) for v in row)
                    if key in seen:
                        dup_count += 1
                    else:
                        seen.add(key)
                if dup_count > 0:
                    issues["重复行"] = [f"  发现 {dup_count} 行重复数据"]
                    report_lines.append(f"[重复行] 发现 {dup_count} 行重复")
                else:
                    report_lines.append("[重复行] 无重复行，通过")

            # Generate report
            report_text = "\n".join(report_lines)
            if not report_text:
                report_text = "所有检查通过"

            # Write report to XLSX
            if issues:
                report_path = os.path.join(self._output_dir, "数据校验报告.xlsx")
                out_wb = Workbook()
                # Remove default sheet
                out_wb.remove(out_wb.active)
                for check_name, check_issues in issues.items():
                    ws = out_wb.create_sheet(title=check_name[:31])
                    ws.append([check_name])
                    ws.append([])
                    for issue in check_issues:
                        ws.append([issue])
                out_wb.save(report_path)
                out_wb.close()
                report_text += f"\n\n详细报告已保存：数据校验报告.xlsx"

            self.progress.emit(total_rows, total_rows, "完成")
            self.finished.emit(report_text)

        except Exception as e:
            self.error_occurred.emit(str(e))
