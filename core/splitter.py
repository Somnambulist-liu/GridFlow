import os
from typing import Dict
import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QThread, Signal

from core.reader import (
    read_sheet_grouped_with_styles,
    read_leading_rows_with_styles,
    copy_merged_cells,
    copy_merged_cells_with_map,
    _apply_style_to_cell,
    _extract_row_styles,
    _copy_column_widths,
    _copy_row_heights,
)


class SplitWorker(QThread):
    progress = Signal(int, int, str)
    finished = Signal(str)
    error_occurred = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._is_cancelled = False

    def cancel(self):
        self._is_cancelled = True

    def configure(
        self,
        file_path: str,
        sheet_name: str,
        column: str,
        mode: str,
        output_dir: str,
        output_path: str,
        name_pattern: str = "{value}.xlsx",
        keep_header: bool = True,
        preserve_formulas: bool = False,
        header_row: int = 1,
        include_lead_rows: bool = True,
        include_tail_rows: bool = False,
        tail_rows_start: int = 0,
        tail_rows_end: int = 0,
    ):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.column = column
        self.mode = mode
        self.output_dir = output_dir
        self.output_path = output_path
        self.name_pattern = name_pattern
        self.keep_header = keep_header
        self.preserve_formulas = preserve_formulas
        self.header_row = header_row
        self.include_lead_rows = include_lead_rows
        self.include_tail_rows = include_tail_rows
        self.tail_rows_start = tail_rows_start
        self.tail_rows_end = tail_rows_end

    def run(self):
        try:
            if self.preserve_formulas:
                self._run_with_formulas()
            else:
                self._run_values_only()
        except Exception as e:
            self.error_occurred.emit(str(e))

    # ── values-only mode (preserves cell styles) ──────────────────

    def _run_values_only(self):
        self.progress.emit(0, 1, "正在读取数据...")
        headers, header_styles, groups = read_sheet_grouped_with_styles(
            self.file_path, self.sheet_name, self.column,
            header_row=self.header_row, data_only=True
        )

        if not groups:
            self.finished.emit("没有数据需要拆分")
            return

        # 从分组中排除尾部行
        groups = self._exclude_tail_from_groups(groups)
        if not groups:
            self.finished.emit("没有数据需要拆分")
            return

        # 读取前置行和尾部行
        leading_rows = self._load_leading_rows()
        tail_rows = self._load_tail_rows()

        os.makedirs(self.output_dir, exist_ok=True)
        total = len(groups)

        if self.mode == "files":
            summary_parts = self._split_to_files(
                headers, header_styles, groups, total, leading_rows=leading_rows, tail_rows=tail_rows)
        else:
            summary_parts = self._split_to_sheets(
                headers, header_styles, groups, total, leading_rows=leading_rows, tail_rows=tail_rows)

        unit = "文件" if self.mode == "files" else "Sheet"
        total_rows = sum(p[1] for p in summary_parts)
        summary = f"拆分完成！共生成 {len(summary_parts)} 个{unit}，总计 {total_rows} 行"
        self.progress.emit(total, total, summary)
        self.finished.emit(summary)

    def _load_leading_rows(self):
        """读取前置行（header_row > 1 且开关开启时自动读取第 1 到 header_row-1 行）。"""
        if self.include_lead_rows and self.header_row > 1:
            return read_leading_rows_with_styles(
                self.file_path, self.sheet_name,
                1, self.header_row - 1, data_only=True
            )
        return []

    def _load_tail_rows(self):
        """读取尾部行（开关开启且有有效范围）。返回 [(src_row, values, styles), ...] 或空列表。"""
        if (self.include_tail_rows and self.tail_rows_start > 0
                and self.tail_rows_end >= self.tail_rows_start):
            return read_leading_rows_with_styles(
                self.file_path, self.sheet_name,
                self.tail_rows_start, self.tail_rows_end, data_only=True
            )
        return []

    def _exclude_tail_from_groups(self, groups: dict) -> dict:
        """从分组数据中移除尾部行（避免尾部行被当作数据拆分）。"""
        if not self.include_tail_rows or self.tail_rows_start <= 0:
            return groups
        tail_src_rows = set(range(self.tail_rows_start, self.tail_rows_end + 1))
        cleaned = {}
        for key, rows in groups.items():
            filtered = [(r, v, s) for r, v, s in rows if r not in tail_src_rows]
            if filtered:
                cleaned[key] = filtered
        return cleaned

    def _extract_leading_from_all_rows(self, all_rows: list):
        """从已加载的 all_rows 列表中提取前置行（公式模式用）。"""
        if not self.include_lead_rows or self.header_row <= 1:
            return []
        if len(all_rows) < self.header_row - 1:
            return []
        leading = []
        for i, row_cells in enumerate(all_rows[:self.header_row - 1], start=1):
            row_values = [c.value for c in row_cells]
            row_styles = _extract_row_styles(row_cells)
            leading.append((i, row_values, row_styles))
        return leading

    def _extract_tail_from_all_rows(self, all_rows: list):
        """从已加载的 all_rows 列表中提取尾部行（公式模式用）。"""
        if not (self.include_tail_rows and self.tail_rows_start > 0
                and self.tail_rows_end >= self.tail_rows_start):
            return []
        if len(all_rows) < self.tail_rows_end:
            return []
        tail = []
        for i, row_cells in enumerate(all_rows[self.tail_rows_start - 1:self.tail_rows_end],
                                      start=self.tail_rows_start):
            row_values = [c.value for c in row_cells]
            row_styles = _extract_row_styles(row_cells)
            tail.append((i, row_values, row_styles))
        return tail

    def _split_to_files(self, headers, header_styles, groups, total, col_letters=None,
                        leading_rows=None, tail_rows=None):
        """按分组值拆分为多个 .xlsx 文件，保留单元格样式和列宽/行高。

        groups 格式: {value: [(src_row, row_values, row_styles), ...]}
        leading_rows: [(src_row, row_values, row_styles), ...] 前置行
        tail_rows: [(src_row, row_values, row_styles), ...] 尾部行
        """
        _leading = leading_rows or []
        _tail = tail_rows or []
        lead_count = len(_leading)
        tail_count = len(_tail)
        summary = []
        src_wb = openpyxl.load_workbook(self.file_path)
        src_ws = src_wb[self.sheet_name]

        for i, (value, rows) in enumerate(groups.items()):
            if self._is_cancelled:
                break
            safe_name = _safe_filename(str(value))
            filename = self.name_pattern.replace("{value}", safe_name)
            filepath = os.path.join(self.output_dir, filename)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = safe_name[:31]

            row_map = {}
            cur = 1

            # 1. 前置行
            for l_idx, (src_row, row_values, row_styles) in enumerate(_leading, cur):
                self._write_row(ws, l_idx, row_values, styles=row_styles)
                row_map[src_row] = l_idx
                cur = l_idx + 1

            # 2. 表头行
            if self.keep_header:
                self._write_row(ws, cur, headers, col_letters, styles=header_styles)
                row_map[self.header_row] = cur
                cur += 1

            # 3. 数据行
            for r_idx, (src_row, row_values, row_styles) in enumerate(rows, cur):
                self._write_row(ws, r_idx, row_values, col_letters, src_row, styles=row_styles)
                row_map[src_row] = r_idx
                cur = r_idx + 1

            # 4. 尾部行
            for t_idx, (src_row, row_values, row_styles) in enumerate(_tail, cur):
                self._write_row(ws, t_idx, row_values, styles=row_styles)
                row_map[src_row] = t_idx
                cur = t_idx + 1

            # 5. 复制列宽和行高
            _copy_column_widths(src_ws, ws)
            _copy_row_heights(src_ws, ws, row_map)

            # 6. 复制合并单元格
            copy_merged_cells_with_map(src_ws, ws, row_map)

            wb.save(filepath)
            wb.close()

            summary.append((value, len(rows)))
            self.progress.emit(i + 1, total, f"正在生成：{safe_name} ({len(rows)} 行)")

        src_wb.close()
        return summary

    def _split_to_sheets(self, headers, header_styles, groups, total, col_letters=None,
                         leading_rows=None, tail_rows=None):
        """按分组值拆分为单个 .xlsx 中的多个 Sheet。

        leading_rows: [(src_row, row_values, row_styles), ...] 前置行
        tail_rows: [(src_row, row_values, row_styles), ...] 尾部行
        """
        _leading = leading_rows or []
        _tail = tail_rows or []
        summary = []
        src_wb = openpyxl.load_workbook(self.file_path)
        src_ws = src_wb[self.sheet_name]

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for i, (value, rows) in enumerate(groups.items()):
            if self._is_cancelled:
                break
            safe_name = _safe_sheet_name(str(value))
            ws = wb.create_sheet(title=safe_name)

            row_map = {}
            cur = 1

            # 1. 前置行
            for l_idx, (src_row, row_values, row_styles) in enumerate(_leading, cur):
                self._write_row(ws, l_idx, row_values, styles=row_styles)
                row_map[src_row] = l_idx
                cur = l_idx + 1

            # 2. 表头行
            if self.keep_header:
                self._write_row(ws, cur, headers, col_letters, styles=header_styles)
                row_map[self.header_row] = cur
                cur += 1

            # 3. 数据行
            for r_idx, (src_row, row_values, row_styles) in enumerate(rows, cur):
                self._write_row(ws, r_idx, row_values, col_letters, src_row, styles=row_styles)
                row_map[src_row] = r_idx
                cur = r_idx + 1

            # 4. 尾部行
            for t_idx, (src_row, row_values, row_styles) in enumerate(_tail, cur):
                self._write_row(ws, t_idx, row_values, styles=row_styles)
                row_map[src_row] = t_idx
                cur = t_idx + 1

            # 5. 复制列宽和行高
            _copy_column_widths(src_ws, ws)
            _copy_row_heights(src_ws, ws, row_map)

            # 6. 复制合并单元格
            copy_merged_cells_with_map(src_ws, ws, row_map)

            summary.append((value, len(rows)))
            self.progress.emit(i + 1, total, f"正在生成 Sheet：{safe_name} ({len(rows)} 行)")

        wb.save(self.output_path)
        wb.close()
        src_wb.close()
        return summary

    # ── formula-preserving mode (preserves formulas AND cell styles) ──

    def _run_with_formulas(self):
        """Split while preserving cell formulas AND cell styles.

        Opens source in normal mode, extracts cell values (including formula
        strings) and styles before closing the source workbook.
        """
        self.progress.emit(0, 1, "正在读取数据（保留公式和样式）...")
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb[self.sheet_name]

        # Read all rows at once, extracting values and styles immediately
        all_rows = list(ws.iter_rows())

        if len(all_rows) < self.header_row:
            wb.close()
            self.finished.emit("表格为空")
            return

        header_cells = all_rows[self.header_row - 1]
        headers = [c.value for c in header_cells]
        header_styles = _extract_row_styles(header_cells)
        headers_str = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(headers)]

        if self.column not in headers_str:
            wb.close()
            self.finished.emit(f"未找到字段：{self.column}")
            return

        col_idx = headers_str.index(self.column)

        # Pre-compute column letters for formula coordinate translation
        col_letters = [get_column_letter(i) for i in range(1, len(header_cells) + 1)]

        # Extract values and styles immediately — do NOT store Cell objects past wb.close()
        # Store (src_row, row_values, row_styles) so formulas can be translated
        groups: Dict[str, list] = {}
        for src_row, row_cells in enumerate(all_rows[self.header_row:], start=self.header_row + 1):
            key = row_cells[col_idx].value if col_idx < len(row_cells) else None
            key = "(空)" if key is None else str(key)
            row_values = [c.value for c in row_cells]
            row_styles = _extract_row_styles(row_cells)
            groups.setdefault(key, []).append((src_row, row_values, row_styles))

        # 提取前置行和尾部行（若配置了有效范围）
        leading_rows = self._extract_leading_from_all_rows(all_rows)
        tail_rows = self._extract_tail_from_all_rows(all_rows)

        wb.close()

        # 从分组中排除尾部行
        groups = self._exclude_tail_from_groups(groups)

        if not groups:
            self.finished.emit("没有数据需要拆分")
            return

        os.makedirs(self.output_dir, exist_ok=True)
        total = len(groups)

        if self.mode == "files":
            summary_parts = self._split_files_formulas(
                headers, header_styles, col_letters, groups, total, leading_rows=leading_rows, tail_rows=tail_rows)
        else:
            summary_parts = self._split_sheets_formulas(
                headers, header_styles, col_letters, groups, total, leading_rows=leading_rows, tail_rows=tail_rows)

        unit = "文件" if self.mode == "files" else "Sheet"
        total_rows = sum(p[1] for p in summary_parts)
        summary = f"拆分完成！共生成 {len(summary_parts)} 个{unit}，总计 {total_rows} 行（含公式）"
        self.progress.emit(total, total, summary)
        self.finished.emit(summary)

    def _split_files_formulas(self, headers, header_styles, col_letters, groups, total,
                              leading_rows=None, tail_rows=None):
        """按分组值拆分为多个文件，保留公式和单元格样式。"""
        _leading = leading_rows or []
        _tail = tail_rows or []
        summary = []
        src_wb = openpyxl.load_workbook(self.file_path)
        src_ws = src_wb[self.sheet_name]

        for i, (value, rows) in enumerate(groups.items()):
            if self._is_cancelled:
                break
            safe_name = _safe_filename(str(value))
            filename = self.name_pattern.replace("{value}", safe_name)
            filepath = os.path.join(self.output_dir, filename)

            out_wb = openpyxl.Workbook()
            out_ws = out_wb.active
            out_ws.title = safe_name[:31]

            row_map = {}
            cur = 1

            # 1. 前置行
            for l_idx, (src_row, row_values, row_styles) in enumerate(_leading, cur):
                self._write_row(out_ws, l_idx, row_values, col_letters, src_row, styles=row_styles)
                row_map[src_row] = l_idx
                cur = l_idx + 1

            # 2. 表头行
            if self.keep_header:
                self._write_row(out_ws, cur, headers, col_letters, styles=header_styles)
                row_map[self.header_row] = cur
                cur += 1

            # 3. 数据行
            for r_idx, (src_row, row_values, row_styles) in enumerate(rows, cur):
                self._write_row(out_ws, r_idx, row_values, col_letters, src_row, styles=row_styles)
                row_map[src_row] = r_idx
                cur = r_idx + 1

            # 4. 尾部行
            for t_idx, (src_row, row_values, row_styles) in enumerate(_tail, cur):
                self._write_row(out_ws, t_idx, row_values, col_letters, src_row, styles=row_styles)
                row_map[src_row] = t_idx
                cur = t_idx + 1

            # 5. 复制列宽和行高
            _copy_column_widths(src_ws, out_ws)
            _copy_row_heights(src_ws, out_ws, row_map)

            # 6. 复制合并单元格
            copy_merged_cells_with_map(src_ws, out_ws, row_map)

            out_wb.save(filepath)
            out_wb.close()

            summary.append((value, len(rows)))
            self.progress.emit(i + 1, total, f"正在生成：{safe_name} ({len(rows)} 行)")

        src_wb.close()
        return summary

    def _split_sheets_formulas(self, headers, header_styles, col_letters, groups, total,
                               leading_rows=None, tail_rows=None):
        """按分组值拆分为单个文件中的多个 Sheet，保留公式和单元格样式。"""
        _leading = leading_rows or []
        _tail = tail_rows or []
        summary = []
        src_wb = openpyxl.load_workbook(self.file_path)
        src_ws = src_wb[self.sheet_name]

        out_wb = openpyxl.Workbook()
        out_wb.remove(out_wb.active)

        for i, (value, rows) in enumerate(groups.items()):
            if self._is_cancelled:
                break
            safe_name = _safe_sheet_name(str(value))
            out_ws = out_wb.create_sheet(title=safe_name)

            row_map = {}
            cur = 1

            # 1. 前置行
            for l_idx, (src_row, row_values, row_styles) in enumerate(_leading, cur):
                self._write_row(out_ws, l_idx, row_values, col_letters, src_row, styles=row_styles)
                row_map[src_row] = l_idx
                cur = l_idx + 1

            # 2. 表头行
            if self.keep_header:
                self._write_row(out_ws, cur, headers, col_letters, styles=header_styles)
                row_map[self.header_row] = cur
                cur += 1

            # 3. 数据行
            for r_idx, (src_row, row_values, row_styles) in enumerate(rows, cur):
                self._write_row(out_ws, r_idx, row_values, col_letters, src_row, styles=row_styles)
                row_map[src_row] = r_idx
                cur = r_idx + 1

            # 4. 尾部行
            for t_idx, (src_row, row_values, row_styles) in enumerate(_tail, cur):
                self._write_row(out_ws, t_idx, row_values, col_letters, src_row, styles=row_styles)
                row_map[src_row] = t_idx
                cur = t_idx + 1

            # 5. 复制列宽和行高
            _copy_column_widths(src_ws, out_ws)
            _copy_row_heights(src_ws, out_ws, row_map)

            # 6. 复制合并单元格
            copy_merged_cells_with_map(src_ws, out_ws, row_map)

            summary.append((value, len(rows)))
            self.progress.emit(i + 1, total, f"正在生成 Sheet：{safe_name} ({len(rows)} 行)")

        out_wb.save(self.output_path)
        out_wb.close()
        src_wb.close()
        return summary

    @staticmethod
    def _write_row(ws, row_idx, values, col_letters=None, src_row=None, styles=None):
        """Write a row of values, translating formula references and applying cell styles.

        Args:
            ws: target worksheet
            row_idx: 1-based row index in the target sheet
            values: list of cell values
            col_letters: column letters for formula translation
            src_row: source row number for formula translation
            styles: list of style dicts (from _extract_cell_style), or None
        """
        for col_idx, val in enumerate(values, 1):
            if col_letters and src_row and isinstance(val, str) and val.startswith("="):
                src_cell = f"{col_letters[col_idx - 1]}{src_row}"
                tgt_cell = f"{get_column_letter(col_idx)}{row_idx}"
                try:
                    val = Translator(val, origin=src_cell).translate_formula(tgt_cell)
                except Exception:
                    pass  # formula references out of range — keep original
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if styles and col_idx - 1 < len(styles):
                _apply_style_to_cell(cell, styles[col_idx - 1])


def _safe_filename(name: str) -> str:
    invalid = '<>:"/\\|?*'
    for ch in invalid:
        name = name.replace(ch, "_")
    return name.strip()[:100]


def _safe_sheet_name(name: str) -> str:
    invalid = '[]:*?/\\'
    for ch in invalid:
        name = name.replace(ch, "_")
    return name.strip()[:31]
