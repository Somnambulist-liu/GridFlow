"""列操作引擎"""
import os
from openpyxl import load_workbook, Workbook
from PySide6.QtCore import QThread, Signal


class ColumnOpsWorker(QThread):
    progress = Signal(int, int, str)
    finished = Signal(str)
    error_occurred = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._file_path = ""
        self._sheet_name = ""
        self._kept_columns = []
        self._renames = {}
        self._order = []
        self._calc_columns = []
        self._output_dir = ""
        self._output_name = ""

    def configure(self, file_path: str, sheet_name: str,
                  kept_columns: list, renames: dict, order: list,
                  calc_columns: list, output_dir: str = "",
                  output_name: str = "列操作结果.xlsx"):
        self._file_path = file_path
        self._sheet_name = sheet_name
        self._kept_columns = kept_columns
        self._renames = renames
        self._order = order
        self._calc_columns = calc_columns
        self._output_dir = output_dir
        self._output_name = output_name

    def run(self):
        try:
            wb = load_workbook(self._file_path, read_only=True)
            ws = wb[self._sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            headers = list(rows[0]) if rows else []
            header_index = {h: i for i, h in enumerate(headers)}

            # Determine final column order
            final_cols = self._order if self._order else self._kept_columns
            final_headers = [self._renames.get(c, c) for c in final_cols]
            for calc in self._calc_columns:
                final_headers.append(calc["name"])

            total_rows = len(rows) - 1
            out_wb = Workbook()
            out_ws = out_wb.active
            out_ws.append(final_headers)

            for row_idx, row in enumerate(rows[1:], 1):
                if row_idx % 500 == 0:
                    self.progress.emit(row_idx, total_rows, f"正在处理 {row_idx}/{total_rows} 行...")

                out_row = []
                for col in final_cols:
                    if col in header_index:
                        out_row.append(row[header_index[col]])
                    else:
                        out_row.append("")

                for calc in self._calc_columns:
                    try:
                        expr = calc["expression"]
                        for col in header_index:
                            expr = expr.replace("{" + col + "}", str(row[header_index[col]] or 0))
                        result = float(eval(expr))
                        out_row.append(round(result, 2) if result != int(result) else int(result))
                    except Exception:
                        out_row.append("")

                out_ws.append(out_row)

            output_path = os.path.join(self._output_dir, self._output_name)
            out_wb.save(output_path)
            out_wb.close()

            self.progress.emit(total_rows, total_rows, "完成")
            self.finished.emit(
                f"列操作完成！{len(final_headers)} 列 × {total_rows} 行\n输出文件：{os.path.basename(output_path)}"
            )

        except Exception as e:
            self.error_occurred.emit(str(e))
