import os
import openpyxl
from PySide6.QtCore import QThread, Signal


class DedupWorker(QThread):
    progress = Signal(int, int, str)
    finished = Signal(str)
    error_occurred = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._is_cancelled = False

    def cancel(self):
        self._is_cancelled = True

    def configure(
        self,
        file_path: str,
        sheet_name: str,
        columns: list,
        keep: str = "first",
        output_dir: str = "",
    ):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.columns = columns
        self.keep = keep
        self.output_dir = output_dir

    def run(self):
        try:
            self.progress.emit(0, 3, "正在读取数据...")
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            ws = wb[self.sheet_name]
            it = ws.iter_rows(values_only=True)
            try:
                headers = [str(c) if c is not None else f"Col{j}" for j, c in enumerate(next(it))]
            except StopIteration:
                wb.close()
                self.finished.emit("没有数据")
                return

            col_indices = [headers.index(c) for c in self.columns if c in headers]
            if not col_indices:
                wb.close()
                self.finished.emit("未找到指定的去重列")
                return

            rows = []
            seen = set()
            dup_count = 0
            for row in it:
                key = tuple(row[i] for i in col_indices if i < len(row))
                if key in seen:
                    dup_count += 1
                    if self.keep == "last":
                        # Replace the earlier row
                        for ri, (r, k) in enumerate(rows):
                            if k == key:
                                rows[ri] = (tuple(row), key)
                                break
                else:
                    seen.add(key)
                    rows.append((tuple(row), key))
            wb.close()

            if dup_count == 0:
                self.finished.emit("没有发现重复数据")
                return

            self.progress.emit(1, 3, f"发现 {dup_count} 行重复，正在生成去重文件...")
            base = os.path.splitext(os.path.basename(self.file_path))[0]
            out_path = os.path.join(self.output_dir, f"{base}_去重结果.xlsx")
            owb = openpyxl.Workbook()
            ows = owb.active
            ows.append(headers)
            for row_data, _ in rows:
                ows.append(row_data)
            owb.save(out_path)
            owb.close()

            summary = f"去重完成！删除了 {dup_count} 行重复数据，保留 {len(rows)} 行"
            self.progress.emit(3, 3, summary)
            self.finished.emit(summary)

        except Exception as e:
            self.error_occurred.emit(str(e))
